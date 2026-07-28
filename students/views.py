from django.db import models
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
                    Font,
                    Alignment,
                    Border,
                    Side
                )
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from .models import (Student,
                     Programme,
                     Department, 
                     AcademicYear, 
                     Semester, 
                     Course, 
                     Unit,
                     Registration,
                     SemesterEnrollment,
                     Applicant,
                     Intake,
                     LecturerAssignment,
                     Result,
                     ResultBatch,
                     ResultBatchLog,
                     ProgrammeLevel,
                     UnitOffering,
                     )
from finance.models import (
    FeeStructure,
    StudentInvoice,
    InvoiceItem,
    FinancialClearance,
)
from finance.services import (
    generate_student_invoice,
    update_financial_clearance,
)
from django.db.models import Q
from django.core.paginator import Paginator
from .utils import generate_admission_no
from django.db import transaction
from django.db.models import Count
from django.contrib.auth.models import User, Group
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import (
    login_required,
    permission_required,
)

from django.urls import reverse_lazy
from django.views.generic import DeleteView
from .forms import (
    DepartmentForm,
    ProgrammeForm,
    StudentForm,
    AcademicYearForm,
    SemesterForm,
    CourseForm,
    UnitForm,
    RegistrationForm,
    SemesterEnrollmentForm,
    ApplicantForm,
    IntakeForm,
    LecturerAssignmentForm,
    ProgrammeLevelForm,
    UnitOfferingForm,
    BulkUnitOfferingForm,
)
from students.services import progress_student as progress_student_service

# Create your views here.
@login_required
def home(request):

    context = {}

    active_year = AcademicYear.objects.filter(
        is_active=True
    ).first()

    active_semester = Semester.objects.filter(
        is_active=True
    ).first()

    context["active_year"] = active_year
    context["active_semester"] = active_semester

    # ==============================
    # Student Dashboard
    # ==============================
    if hasattr(request.user, "student_profile"):

        student = request.user.student_profile

        enrollment = (
            SemesterEnrollment.objects
            .select_related(
                "programme_level",
                "programme_level__programme",
                "academic_year",
                "semester",
            )
            .filter(student=student)
            .order_by("-academic_year__id", "-semester__id")
            .first()
        )

        registrations = Registration.objects.none()

        registration_count = 0

        if enrollment:

            registrations = (
                Registration.objects
                .select_related("unit")
                .filter(
                    enrollment=enrollment,
                    status=Registration.REGISTERED,
                )
                .order_by("unit__code")
            )

            registration_count = registrations.count()

        context.update({

            "dashboard_type": "student",

            "student": student,

            "enrollment": enrollment,

            "registrations": registrations,

            "registration_count": registration_count,

        })

    # ==============================
    # Administrator Dashboard
    # ==============================

    elif request.user.is_superuser:

        context.update({

            "dashboard_type": "admin",

            "total_students": Student.objects.count(),

            "total_programmes": Programme.objects.count(),

            "total_departments": Department.objects.count(),

            "total_applicants": Applicant.objects.count(),

            "total_enrollments": SemesterEnrollment.objects.count(),

            "total_registrations": Registration.objects.count(),

            "total_users": User.objects.count(),

        })

    # ==============================
    # Lecturer Dashboard
    # ==============================
    elif request.user.groups.filter(
        name="Lecturer"
    ).exists():

        assignments = LecturerAssignment.objects.filter(
            lecturer=request.user,
            unit_offering__is_active=True,
            unit_offering__academic_year__is_active=True,
            unit_offering__semester__is_active=True,
        )

        my_units = assignments.count()


        my_students = Registration.objects.filter(
            unit__in=assignments.values(
                "unit_offering__unit"
            ),
            enrollment__academic_year__is_active=True,
            enrollment__semester__is_active=True,
            status=Registration.REGISTERED,
        ).values(
            "enrollment__student"
        ).distinct().count()


        pending_results = Result.objects.filter(
            unit_offering__in=assignments.values(
                "unit_offering"
            ),
            batch__status=ResultBatch.SUBMITTED,
        ).count()


        context.update({

            "dashboard_type": "lecturer",

            "my_units": my_units,

            "my_students": my_students,

            "pending_results": pending_results,

        })
    # ==============================
    # Exam Officer Dashboard
    # ==============================
    elif request.user.has_perm(
        "students.view_lecturerassignment"
    ):

        submitted_batches = ResultBatch.objects.filter(
            status=ResultBatch.SUBMITTED
        )

        returned_batches = ResultBatch.objects.filter(
            status=ResultBatch.RETURNED
        )

        approved_batches = ResultBatch.objects.filter(
            status=ResultBatch.APPROVED
        )

        unlocked_batches = ResultBatch.objects.filter(
            status=ResultBatch.UNLOCKED
        )

        context.update({

            "dashboard_type": "exam",

            "submitted_batches": submitted_batches.count(),

            "returned_batches": returned_batches.count(),

            "approved_batches": approved_batches.count(),

            "unlocked_batches": unlocked_batches.count(),

            "total_batches": ResultBatch.objects.count(),

            "total_results": Result.objects.count(),

            "total_assignments": LecturerAssignment.objects.count(),

            "total_unit_offerings": UnitOffering.objects.count(),

        })

    # ==============================
    # Registrar Dashboard
    # ==============================

    elif request.user.has_perm(
        "students.view_registration"
    ):

        context.update({

            "dashboard_type": "registrar",

            "total_students": Student.objects.count(),

            "total_enrollments": SemesterEnrollment.objects.count(),

            "total_registrations": Registration.objects.count(),

        })

    # ==============================
    # Admissions Dashboard
    # ==============================

    elif request.user.has_perm(
        "students.view_applicant"
    ):

        context.update({

            "dashboard_type": "admissions",

            "total_applicants":

                Applicant.objects.count(),

            "pending_applicants":

                Applicant.objects.filter(
                    status="PENDING"
                ).count(),

            "approved_applicants":

                Applicant.objects.filter(
                    status="APPROVED"
                ).count(),

            "rejected_applicants":

                Applicant.objects.filter(
                    status="REJECTED"
                ).count(),

            "total_intakes":

                Intake.objects.count(),

        })

    else:

        context["dashboard_type"] = "general"

    latest_enrollment = None

    if hasattr(request.user, "student_profile"):
        latest_enrollment = (
            request.user.student_profile.enrollments
            .select_related("academic_year", "semester")
            .order_by("-academic_year__id", "-semester__id")
            .first()
        )

    context["latest_enrollment"] = latest_enrollment

    return render(
        request,
        "students/home.html",
        context,
    )


@login_required
@permission_required(
    "students.view_lecturerassignment",
    raise_exception=True,
)
def exam_dashboard(request):

    submitted_batches = (
        ResultBatch.objects
        .select_related(
            "unit_offering",
            "unit_offering__unit",
            "unit_offering__programme_level",
            "lecturer_assignment",
            "lecturer_assignment__lecturer",
        )
        .filter(
            status=ResultBatch.SUBMITTED
        )
        .order_by(
            "-submitted_at"
        )
    )

    returned_batches = (
        ResultBatch.objects.filter(
            status=ResultBatch.RETURNED
        ).count()
    )

    approved_batches = (
        ResultBatch.objects.filter(
            status=ResultBatch.APPROVED
        ).count()
    )

    unlocked_batches = (
        ResultBatch.objects.filter(
            status=ResultBatch.UNLOCKED
        ).count()
    )

    draft_batches = (
        ResultBatch.objects.filter(
            status=ResultBatch.DRAFT
        ).count()
    )

    context = {

        "submitted_batches": submitted_batches,

        "submitted_count": submitted_batches.count(),

        "returned_count": returned_batches,

        "approved_count": approved_batches,

        "unlocked_count": unlocked_batches,

        "draft_count": draft_batches,

        "total_batches": ResultBatch.objects.count(),

        "total_results": Result.objects.count(),

        "total_assignments": LecturerAssignment.objects.count(),

        "total_unit_offerings": UnitOffering.objects.count(),

    }

    return render(
        request,
        "students/exams/exam_dashboard.html",
        context,
    )

@login_required
@permission_required(
    "students.view_student",
    raise_exception=True,
)
def student_list(request):

    query = request.GET.get("q")


    students = (
        Student.objects
        .select_related(
            "programme",
            "programme__course",
            "programme__course__department",
        )
        .order_by(
            "admission_no"
        )
    )


    if query:

        students = students.filter(
            Q(admission_no__icontains=query)
            |
            Q(first_name__icontains=query)
            |
            Q(last_name__icontains=query)
        )


    paginator = Paginator(
        students,
        10
    )


    page_number = request.GET.get(
        "page"
    )


    page_obj = paginator.get_page(
        page_number
    )


    return render(
        request,
        "students/student_list.html",
        {
            "page_obj": page_obj,
            "query": query,
            "total_students": students.count(),
        }
    )

@login_required
@permission_required(
    "students.add_student",
    raise_exception=True,
)
def student_create(request):

    if request.method == "POST":

        form = StudentForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student created successfully."
            )

            return redirect(
                "student_list"
            )

    else:

        form = StudentForm()


    return render(
        request,
        "students/student_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.view_student",
    raise_exception=True,
)

@login_required
@permission_required(
    "students.view_student",
    raise_exception=True
)
def student_detail(request, id):

    student = get_object_or_404(
        Student,
        id=id
    )


    enrollments = (
        student.enrollments
        .select_related(
            "academic_year",
            "semester",
            "programme",
            "programme_level",
        )
        .order_by(
            "-academic_year",
            "-semester"
        )
    )


    return render(
        request,
        "students/student_detail.html",
        {
            "student": student,
            "enrollments": enrollments,
        }
    )

@login_required
@permission_required(
    "students.change_student",
    raise_exception=True,
)
def student_update(request, id):

    student = get_object_or_404(
        Student,
        id=id
    )


    if request.method == "POST":

        form = StudentForm(
            request.POST,
            instance=student
        )


        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student updated successfully."
            )

            return redirect(
                "student_detail",
                id=student.id
            )

    else:

        form = StudentForm(
            instance=student
        )


    return render(
        request,
        "students/student_form.html",
        {
            "form": form,
            "student": student,
        }
    )

@login_required
@permission_required(
    "students.delete_student",
    raise_exception=True,
)
def student_delete(request,id):

    student = get_object_or_404(
        Student,
        id=id
    )


    if request.method == "POST":

        student.delete()

        messages.success(
            request,
            "Student deleted successfully."
        )

        return redirect(
            "student_list"
        )


    return render(
        request,
        "students/student_confirm_delete.html",
        {
            "student":student
        }
    )


@login_required
@permission_required(
    "students.view_department",
    raise_exception=True
)
def department_list(request):

    departments = (
        Department.objects
        .all()
        .order_by(
            "name"
        )
    )


    return render(
        request,
        "students/departments/department_list.html",
        {
            "departments": departments
        }
    )

@login_required
@permission_required(
    "students.add_department",
    raise_exception=True
)
def department_create(request):

    if request.method == "POST":

        form = DepartmentForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Department created successfully."
            )

            return redirect("department_list")

    else:

        form = DepartmentForm()

    context = {
        "form": form
    }

    return render(
        request,
        "students/departments/department_form.html",
        context
    )

@login_required
@permission_required(
    "students.change_department",
    raise_exception=True
)
def department_update(request, pk):

    department = get_object_or_404(
        Department,
        pk=pk
    )

    if request.method == "POST":

        form = DepartmentForm(
            request.POST,
            instance=department
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Department updated successfully."
            )

            return redirect("department_list")

    else:

        form = DepartmentForm(
            instance=department
        )

    context = {
        "form": form,
        "department": department
    }

    return render(
        request,
        "students/departments/department_form.html",
        context
    )

@login_required
@permission_required(
    "students.delete_department",
    raise_exception=True
)
def department_delete(request, pk):

    department = get_object_or_404(
        Department,
        pk=pk
    )

    if request.method == "POST":

        department.delete()

        messages.success(
            request,
            "Department deleted successfully."
        )

        return redirect("department_list")

    context = {
        "department": department
    }

    return render(
        request,
        "students/departments/department_confirm_delete.html",
        context
    )

@login_required
@permission_required(
    "students.view_programme",
    raise_exception=True
)
def programme_list(request):

    programmes = (
        Programme.objects
        .select_related(
            "course",
            "course__department",
        )
        .order_by(
            "name"
        )
    )

    context = {
        "programmes": programmes
    }

    return render(
        request,
        "students/programmes/programme_list.html",
        context
    )

@login_required
@permission_required(
    "students.add_programme",
    raise_exception=True
)
def programme_create(request):

    if request.method == "POST":

        form = ProgrammeForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Programme created successfully."
            )

            return redirect(
                "programme_list"
            )

    else:

        form = ProgrammeForm()

    return render(
        request,
        "students/programmes/programme_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_programme",
    raise_exception=True
)
def programme_update(request, pk):

    programme = get_object_or_404(
        Programme,
        pk=pk
    )

    if request.method == "POST":

        form = ProgrammeForm(
            request.POST,
            instance=programme
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Programme updated successfully."
            )

            return redirect(
                "programme_list"
            )

    else:

        form = ProgrammeForm(
            instance=programme
        )

    return render(
        request,
        "students/programmes/programme_form.html",
        {
            "form": form,
            "programme": programme
        }
    )

@login_required
@permission_required(
    "students.delete_programme",
    raise_exception=True
)
def programme_delete(request, pk):

    programme = get_object_or_404(
        Programme,
        pk=pk
    )

    if request.method == "POST":

        programme.delete()

        messages.success(
            request,
            "Programme deleted successfully."
        )

        return redirect(
            "programme_list"
        )

    return render(
        request,
        "students/programmes/programme_confirm_delete.html",
        {
            "programme": programme
        }
    )

@login_required
@permission_required(
    "students.view_academicyear",
    raise_exception=True
)
def academic_year_list(request):

    years = AcademicYear.objects.all()

    return render(
        request,
        "students/academic_years/academic_year_list.html",
        {
            "years": years
        }
    )

@login_required
@permission_required(
    "students.add_academicyear",
    raise_exception=True
)
def academic_year_create(request):

    if request.method == "POST":

        form = AcademicYearForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Academic year created successfully."
            )

            return redirect(
                "academic_year_list"
            )

    else:

        form = AcademicYearForm()

    return render(
        request,
        "students/academic_years/academic_year_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_academicyear",
    raise_exception=True
)
def academic_year_update(request, pk):

    year = get_object_or_404(
        AcademicYear,
        pk=pk
    )

    if request.method == "POST":

        form = AcademicYearForm(
            request.POST,
            instance=year
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Academic year updated successfully."
            )

            return redirect(
                "academic_year_list"
            )

    else:

        form = AcademicYearForm(
            instance=year
        )

    return render(
        request,
        "students/academic_years/academic_year_form.html",
        {
            "form": form
        }
    )


@login_required
@permission_required(
    "students.delete_academicyear",
    raise_exception=True
)
def academic_year_delete(request, pk):

    year = get_object_or_404(
        AcademicYear,
        pk=pk
    )

    if request.method == "POST":

        year.delete()

        messages.success(
            request,
            "Academic year deleted successfully."
        )

        return redirect(
            "academic_year_list"
        )

    return render(
        request,
        "students/academic_years/academic_year_confirm_delete.html",
        {
            "year": year
        }
    )


@login_required
@permission_required(
    "students.change_academicyear",
    raise_exception=True
)
def open_registration(request, pk):

    year = get_object_or_404(
        AcademicYear,
        pk=pk
    )

    year.registration_open = True

    year.save()

    messages.success(
        request,
        f"Registration opened for {year.year_name}"
    )

    return redirect(
        "academic_year_list"
    )


@login_required
@permission_required(
    "students.change_academicyear",
    raise_exception=True
)
def close_registration(request, pk):

    year = get_object_or_404(
        AcademicYear,
        pk=pk
    )

    year.registration_open = False

    year.save()

    messages.success(
        request,
        f"Registration closed for {year.year_name}"
    )

    return redirect(
        "academic_year_list"
    )

@login_required
@permission_required(
    "students.view_semester",
    raise_exception=True
)
def semester_list(request):

    semesters = Semester.objects.select_related(
        "academic_year"
    )

    return render(
        request,
        "students/semesters/semester_list.html",
        {
            "semesters": semesters
        }
    )


@login_required
@permission_required(
    "students.add_semester",
    raise_exception=True
)
def semester_create(request):

    if request.method == "POST":

        form = SemesterForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Semester created successfully."
            )

            return redirect(
                "semester_list"
            )

    else:

        form = SemesterForm()

    return render(
        request,
        "students/semesters/semester_form.html",
        {
            "form": form,
            "title": "Create Semester"
        }
    )


@login_required
@permission_required(
    "students.change_semester",
    raise_exception=True
)
def semester_update(request, pk):

    semester = get_object_or_404(
        Semester,
        pk=pk
    )

    if request.method == "POST":

        form = SemesterForm(
            request.POST,
            instance=semester
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Semester updated successfully."
            )

            return redirect(
                "semester_list"
            )

    else:

        form = SemesterForm(
            instance=semester
        )

    return render(
        request,
        "students/semesters/semester_form.html",
        {
            "form": form,
            "title": "Edit Semester"
        }
    )


@login_required
@permission_required(
    "students.delete_semester",
    raise_exception=True
)
def semester_delete(request, pk):

    semester = get_object_or_404(
        Semester,
        pk=pk
    )

    if request.method == "POST":

        semester.delete()

        messages.success(
            request,
            "Semester deleted successfully."
        )

        return redirect(
            "semester_list"
        )

    return render(
        request,
        "students/semesters/semester_confirm_delete.html",
        {
            "semester": semester
        }
    )

@login_required
@permission_required(
    "students.change_semester",
    raise_exception=True
)
def activate_semester(request, pk):

    semester = get_object_or_404(
        Semester,
        pk=pk
    )

    academic_year = (
        semester.academic_year
    )

    academic_year.is_active = True
    academic_year.save()

    semester.is_active = True
    semester.save()

    messages.success(
        request,
        (
            f"{semester.semester_name} "
            f"activated for "
            f"{academic_year.year_name}"
        )
    )

    return redirect(
        "semester_list"
    )


@login_required
@permission_required(
    "students.view_course",
    raise_exception=True
)
def course_list(request):

    courses = (
        Course.objects
        .select_related(
            "department"
        )
        .order_by(
            "name"
        )
    )


    search = request.GET.get(
        "search"
    )


    department_id = request.GET.get(
        "department"
    )


    if search:

        courses = courses.filter(
            name__icontains=search
        )


    if department_id:

        courses = courses.filter(
            department_id=department_id
        )


    context = {

        "courses": courses,

        "departments": (
            Department.objects
            .order_by(
                "name"
            )
        ),

        "search": search,

        "selected_department": department_id,

    }


    return render(
        request,
        "students/courses/course_list.html",
        context
    )


@login_required
@permission_required(
    "students.add_course",
    raise_exception=True
)
def course_create(request):

    if request.method == "POST":

        form = CourseForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Course created successfully."
            )

            return redirect(
                "course_list"
            )

    else:

        form = CourseForm()

    return render(
        request,
        "students/courses/course_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_course",
    raise_exception=True
)
def course_update(request, pk):

    course = get_object_or_404(
        Course,
        pk=pk
    )

    if request.method == "POST":

        form = CourseForm(
            request.POST,
            instance=course
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Course updated successfully."
            )

            return redirect(
                "course_list"
            )

    else:

        form = CourseForm(
            instance=course
        )

    return render(
        request,
        "students/courses/course_form.html",
        {
            "form": form,
            "course": course
        }
    )

@login_required
@permission_required(
    "students.delete_course",
    raise_exception=True
)
def course_delete(request, pk):

    course = get_object_or_404(
        Course,
        pk=pk
    )

    if request.method == "POST":

        course.delete()

        messages.success(
            request,
            "Course deleted successfully."
        )

        return redirect(
            "course_list"
        )

    return render(
        request,
        "students/courses/course_confirm_delete.html",
        {
            "course": course
        }
    )

@login_required
@permission_required(
    "students.view_unit",
    raise_exception=True
)
def unit_list(request):

    units = (
        Unit.objects
        .select_related(
            "programme_level",
            "programme_level__programme",
            "programme_level__programme__course",
        )
        .order_by(
            "programme_level__programme__name",
            "programme_level__progression_order",
            "code",
        )
    )


    search = request.GET.get(
        "search"
    )


    programme_level_id = request.GET.get(
        "programme_level"
    )


    if search:

        units = units.filter(
            name__icontains=search
        )


    if programme_level_id:

        units = units.filter(
            programme_level_id=programme_level_id
        )


    context = {

        "units": units,


        "programme_levels": (
            ProgrammeLevel.objects
            .select_related(
                "programme",
            )
            .filter(
                is_active=True
            )
            .order_by(
                "programme__name",
                "progression_order",
            )
        ),


        "search": search,


        "selected_programme_level":
            programme_level_id,
    }


    return render(
        request,
        "students/units/unit_list.html",
        context
    )

@login_required
@permission_required(
    "students.add_unit",
    raise_exception=True
)
def unit_create(request):

    if request.method == "POST":

        form = UnitForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Unit created successfully."
            )

            return redirect(
                "unit_list"
            )

    else:

        form = UnitForm()

    return render(
        request,
        "students/units/unit_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_unit",
    raise_exception=True
)
def unit_update(request, pk):

    unit = get_object_or_404(
        Unit,
        pk=pk
    )

    if request.method == "POST":

        form = UnitForm(
            request.POST,
            instance=unit
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Unit updated successfully."
            )

            return redirect(
                "unit_list"
            )

    else:

        form = UnitForm(
            instance=unit
        )

    return render(
        request,
        "students/units/unit_form.html",
        {
            "form": form,
            "unit": unit
        }
    )

@login_required
@permission_required(
    "students.delete_unit",
    raise_exception=True
)
def unit_delete(request, pk):

    unit = get_object_or_404(
        Unit,
        pk=pk
    )

    if request.method == "POST":

        unit.delete()

        messages.success(
            request,
            "Unit deleted successfully."
        )

        return redirect(
            "unit_list"
        )

    return render(
        request,
        "students/units/unit_confirm_delete.html",
        {
            "unit": unit
        }
    )


@login_required
@permission_required(
    "students.view_unitoffering",
    raise_exception=True
)
def unit_offering_list(request):

    offerings = (
        UnitOffering.objects
        .select_related(
            "academic_year",
            "semester",
            "programme_level",
            "programme_level__programme",
            "unit",
        )
        .order_by(
            "-academic_year__year_name",
            "semester__semester_name",
            "programme_level__programme__name",
            "programme_level__progression_order",
            "unit__code",
        )
    )


    return render(
        request,
        "students/unit_offerings/unit_offering_list.html",
        {
            "offerings": offerings
        }
    )


@login_required
@permission_required(
    "students.add_unitoffering",
    raise_exception=True
)
def unit_offering_create(request):

    if request.method == "POST":

        form = UnitOfferingForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Unit offering created successfully."
            )

            return redirect(
                "unit_offering_list"
            )

    else:

        form = UnitOfferingForm()


    return render(
        request,
        "students/unit_offerings/unit_offering_form.html",
        {
            "form": form,
            "title": "Create Unit Offering",
        }
    )


@login_required
@permission_required(
    "students.delete_unitoffering",
    raise_exception=True
)
def unit_offering_delete(request, pk):

    offering = get_object_or_404(
        UnitOffering,
        pk=pk
    )


    if request.method == "POST":

        offering.delete()

        messages.success(
            request,
            "Unit offering deleted successfully."
        )

        return redirect(
            "unit_offering_list"
        )


    return render(
        request,
        "students/unit_offerings/unit_offering_confirm_delete.html",
        {
            "offering": offering
        }
    )


@login_required
@permission_required(
    "students.add_unitoffering",
    raise_exception=True,
)
def bulk_unit_offering(request):

    if request.method == "POST":

        form = BulkUnitOfferingForm(request.POST)

        if form.is_valid():

            academic_year = form.cleaned_data["academic_year"]
            semester = form.cleaned_data["semester"]
            programme_level = form.cleaned_data["programme_level"]

            units = Unit.objects.filter(
                programme_level=programme_level,
                is_active=True,
            )

            created = 0
            skipped = 0

            for unit in units:

                _, was_created = UnitOffering.objects.get_or_create(
                    academic_year=academic_year,
                    semester=semester,
                    programme_level=programme_level,
                    unit=unit,
                )

                if was_created:
                    created += 1
                else:
                    skipped += 1

            messages.success(
                request,
                f"{created} unit offering(s) created. "
                f"{skipped} already existed.",
            )

            return redirect("unit_offering_list")

    else:

        form = BulkUnitOfferingForm()

    return render(
        request,
        "students/unit_offerings/bulk_unit_offering.html",
        {
            "form": form,
            "title": "Bulk Unit Offering",
        },
    )

@login_required
@permission_required(
    "students.view_registration",
    raise_exception=True
)
def registration_list(request):

    registrations = (
        Registration.objects
        .select_related(
            "enrollment",
            "enrollment__student",
            "enrollment__programme",
            "enrollment__programme_level",
            "enrollment__academic_year",
            "enrollment__semester",
            "unit",
        )
        .order_by(
            "-enrollment__academic_year",
            "-enrollment__semester",
            "unit__code",
        )
    )


    search = request.GET.get(
        "search"
    )


    if search:

        registrations = registrations.filter(
            Q(
                enrollment__student__admission_no__icontains=search
            )
            |
            Q(
                enrollment__student__first_name__icontains=search
            )
            |
            Q(
                unit__code__icontains=search
            )
        )


    context = {

        "registrations": registrations,

        "search": search,

    }


    return render(
        request,
        "students/registrations/registration_list.html",
        context
    )

@login_required
@permission_required(
    "students.add_registration",
    raise_exception=True
)
def registration_create(request):

    if request.method == "POST":

        form = RegistrationForm(
            request.POST
        )

        if form.is_valid():

            registration = form.save()

            messages.success(
                request,
                "Unit registered successfully."
            )

            return redirect(
                "registration_list"
            )

    else:

        form = RegistrationForm()


    return render(
        request,
        "students/registrations/registration_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_registration",
    raise_exception=True
)
def registration_update(request, pk):

    registration = get_object_or_404(
        Registration,
        pk=pk
    )


    if request.method == "POST":

        form = RegistrationForm(
            request.POST,
            instance=registration
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Registration updated successfully."
            )

            return redirect(
                "registration_list"
            )

    else:

        form = RegistrationForm(
            instance=registration
        )


    return render(
        request,
        "students/registrations/registration_form.html",
        {
            "form": form,
            "registration": registration,
        }
    )

@login_required
@permission_required(
    "students.delete_registration",
    raise_exception=True
)
def registration_delete(request, pk):

    registration = get_object_or_404(
        Registration,
        pk=pk
    )


    if request.method == "POST":

        registration.delete()

        messages.success(
            request,
            "Registration deleted successfully."
        )

        return redirect(
            "registration_list"
        )


    return render(
        request,
        "students/registrations/registration_confirm_delete.html",
        {
            "registration": registration
        }
    )

@login_required
def my_registrations(request):

    student = get_object_or_404(
        Student,
        user=request.user
    )

    registrations = (
        Registration.objects
        .filter(
            enrollment__student=student
        )
        .select_related(
            "enrollment",
            "enrollment__academic_year",
            "enrollment__semester",
            "unit",
        )
        .order_by(
            "-enrollment__academic_year",
            "-enrollment__semester",
            "unit__code",
        )
    )

    active_count = registrations.filter(
        status=Registration.REGISTERED
    ).count()

    dropped_count = registrations.filter(
        status=Registration.DROPPED
    ).count()

    return render(
        request,
        "students/registrations/my_registrations.html",
        {
            "student": student,
            "registrations": registrations,
            "active_count": active_count,
            "dropped_count": dropped_count,
        }
    )


@login_required
def register_units(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment.objects.select_related(
            "student",
            "programme_level",
            "academic_year",
            "semester",
        ),
        pk=pk,
    )

    is_student = hasattr(request.user, "student_profile")

    # ----------------------------------------
    # Student must be enrolled
    # ----------------------------------------

    if enrollment.status != SemesterEnrollment.ENROLLED:

        messages.error(
            request,
            "Only enrolled students can register units."
        )

        if is_student:
            return redirect("home")

        return redirect(
            "semester_enrollment_detail",
            pk=enrollment.pk,
        )

    # ----------------------------------------
    # Registration window must be open
    # ----------------------------------------

    if not enrollment.academic_year.registration_open:

        messages.error(
            request,
            "Unit registration is currently closed."
        )

        if is_student:
            return redirect("home")

        return redirect(
            "semester_enrollment_detail",
            pk=enrollment.pk,
        )

    # ----------------------------------------
    # Financial eligibility
    # ----------------------------------------

    clearance = getattr(
        enrollment,
        "financial_clearance",
        None,
    )

    if not clearance:

        messages.error(
            request,
            "Financial clearance has not been processed for this semester."
        )

        if is_student:
            return redirect("home")

        return redirect(
            "semester_enrollment_detail",
            pk=enrollment.pk,
        )

    if not clearance.registration_cleared:

        messages.error(
            request,
            "You have not met the financial requirements for unit registration."
        )

        if is_student:
            return redirect("home")

        return redirect(
            "semester_enrollment_detail",
            pk=enrollment.pk,
        )

    # ----------------------------------------
    # Units offered
    # ----------------------------------------

    available_units = (
        UnitOffering.objects.filter(
            academic_year=enrollment.academic_year,
            semester=enrollment.semester,
            programme_level=enrollment.programme_level,
            is_active=True,
        )
        .select_related(
            "unit",
        )
        .order_by(
            "unit__code",
        )
    )

    # ----------------------------------------
    # Save Registration
    # ----------------------------------------

    if request.method == "POST":

        unit_ids = request.POST.getlist("units")

        if not unit_ids:

            messages.error(
                request,
                "Please select at least one unit before registering."
            )

            return redirect(
                "register_units",
                pk=enrollment.pk,
            )

        registered = 0
        duplicates = 0

        with transaction.atomic():

            selected_offerings = available_units.filter(
                unit_id__in=unit_ids
            )

            for offering in selected_offerings:

                _, created = Registration.objects.get_or_create(
                    enrollment=enrollment,
                    unit=offering.unit,
                    defaults={
                        "registration_type": Registration.NORMAL,
                    },
                )

                if created:
                    registered += 1
                else:
                    duplicates += 1

        if registered:

            messages.success(
                request,
                f"{registered} unit(s) registered successfully."
            )

        elif duplicates:

            messages.info(
                request,
                "The selected units are already registered."
            )

        # ----------------------------------------
        # Redirect after registration
        # ----------------------------------------

        if is_student:
            return redirect("my_registrations")

        return redirect(
            "semester_enrollment_detail",
            pk=enrollment.pk,
        )

    # ----------------------------------------
    # Display registration page
    # ----------------------------------------

    return render(
        request,
        "students/registrations/register_units.html",
        {
            "enrollment": enrollment,
            "units": available_units,
        },
    )


@login_required
def drop_registration(request, pk):

    student = get_object_or_404(
        Student,
        user=request.user
    )

    registration = get_object_or_404(
        Registration,
        pk=pk,
        enrollment__student=student,
    )

    # Registration already dropped
    if registration.status == Registration.DROPPED:

        messages.warning(
            request,
            "This unit has already been dropped."
        )

        return redirect(
            "my_registrations"
        )

    # Registration lock
    if not registration.enrollment.academic_year.registration_open:

        messages.error(
            request,
            "Registration changes are closed."
        )

        return redirect(
            "my_registrations"
        )

    if request.method == "POST":

        reason = request.POST.get(
            "reason",
            ""
        ).strip()

        registration.status = Registration.DROPPED
        registration.dropped_by = request.user
        registration.dropped_at = timezone.now()
        registration.drop_reason = reason

        registration.save()

        messages.success(
            request,
            "Unit dropped successfully."
        )

        return redirect(
            "my_registrations"
        )

    return render(
        request,
        "students/registrations/drop_registration.html",
        {
            "registration": registration
        }
    )


@login_required
@permission_required(
    "students.view_semesterenrollment",
    raise_exception=True
)
def enrollment_list(request):

    enrollments = (
        SemesterEnrollment.objects
        .select_related(
            "student",
            "programme",
            "programme_level",
            "academic_year",
            "semester",
        )
        .order_by(
            "-academic_year",
            "-semester",
            "student__admission_no"
        )
    )


    # Search

    search = request.GET.get(
        "search"
    )

    if search:

        enrollments = enrollments.filter(

            Q(
                student__admission_no__icontains=search
            )
            |
            Q(
                student__first_name__icontains=search
            )
            |
            Q(
                student__last_name__icontains=search
            )

        )


    # Filters

    academic_year = request.GET.get(
        "academic_year"
    )

    if academic_year:

        enrollments = enrollments.filter(
            academic_year_id=academic_year
        )


    semester = request.GET.get(
        "semester"
    )

    if semester:

        enrollments = enrollments.filter(
            semester_id=semester
        )


    programme = request.GET.get(
        "programme"
    )

    if programme:

        enrollments = enrollments.filter(
            programme_id=programme
        )


    programme_level = request.GET.get(
        "programme_level"
    )

    if programme_level:

        enrollments = enrollments.filter(
            programme_level_id=programme_level
        )


    status = request.GET.get(
        "status"
    )

    if status:

        enrollments = enrollments.filter(
            status=status
        )


    context = {

        "enrollments": enrollments,

        "academic_years":
            AcademicYear.objects.all(),

        "semesters":
            Semester.objects.all(),

        "programmes":
            Programme.objects.all(),

        "programme_levels":
            ProgrammeLevel.objects.all(),

        "statuses":
            SemesterEnrollment.STATUS_CHOICES,

    }


    return render(
        request,
        "students/enrollments/enrollment_list.html",
        context
    )



@login_required
@permission_required(
    "students.add_semesterenrollment",
    raise_exception=True
)
def enrollment_create(request):

    active_year = (
        AcademicYear.objects
        .filter(is_active=True)
        .first()
    )

    active_semester = (
        Semester.objects
        .filter(is_active=True)
        .first()
    )


    if not active_year:

        messages.error(
            request,
            "No active academic year found."
        )

        return redirect(
            "enrollment_list"
        )


    if not active_semester:

        messages.error(
            request,
            "No active semester found."
        )

        return redirect(
            "enrollment_list"
        )


    if request.method == "POST":

        form = SemesterEnrollmentForm(
            request.POST
        )


        if form.is_valid():

            enrollment = form.save(
                commit=False
            )


            # Prevent duplicate enrollment

            exists = (
                SemesterEnrollment.objects
                .filter(
                    student=enrollment.student,
                    academic_year=active_year,
                    semester=active_semester
                )
                .exists()
            )


            if exists:

                messages.error(
                    request,
                    "Student already has an enrollment for this semester."
                )

                return redirect(
                    "enrollment_create"
                )


            # Automatically assign current academic period

            enrollment.academic_year = (
                active_year
            )

            enrollment.semester = (
                active_semester
            )


            # Ensure programme comes from student

            enrollment.programme = (
                enrollment.student.programme
            )


            # Ensure the selected programme level belongs
            # to the student's programme.

            if (
                enrollment.programme_level.programme
                != enrollment.programme
            ):

                messages.error(
                    request,
                    "Selected programme level does not belong to the student's programme."
                )

                return redirect(
                    "enrollment_create"
                )

            enrollment.status = (
                SemesterEnrollment.ENROLLED
            )


            enrollment.save()


            messages.success(
                request,
                "Student enrolled successfully."
            )


            return redirect(
                "enrollment_list"
            )


    else:

        form = SemesterEnrollmentForm()


    return render(
        request,
        "students/enrollments/enrollment_form.html",
        {
            "form": form,
            "active_year": active_year,
            "active_semester": active_semester,
        }
    )



@login_required
@permission_required(
    "students.view_semesterenrollment",
    raise_exception=True
)
def enrollment_detail(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )


    registrations = (
        enrollment.registrations
        .select_related(
            "unit",
            "unit__programme_level"
        )
    )


    return render(
        request,
        "students/enrollments/enrollment_detail.html",
        {
            "enrollment": enrollment,
            "registrations": registrations,
        }
    )



@login_required
@permission_required(
    "students.change_semesterenrollment",
    raise_exception=True
)
def enrollment_update(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )


    if request.method == "POST":


        form = SemesterEnrollmentForm(
            request.POST,
            instance=enrollment
        )


        if form.is_valid():


            form.save()


            messages.success(
                request,
                "Enrollment updated successfully."
            )


            return redirect(
                "enrollment_list"
            )


    else:


        form = SemesterEnrollmentForm(
            instance=enrollment
        )



    return render(
        request,
        "students/enrollments/enrollment_form.html",
        {
            "form": form
        }
    )



@login_required
@permission_required(
    "students.delete_semesterenrollment",
    raise_exception=True
)
def enrollment_delete(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )


    if request.method == "POST":


        enrollment.delete()


        messages.success(
            request,
            "Enrollment deleted successfully."
        )


        return redirect(
            "enrollment_list"
        )



    return render(
        request,
        "students/enrollments/enrollment_confirm_delete.html",
        {
            "enrollment": enrollment
        }
    )

@login_required
@permission_required(
    "students.view_applicant",
    raise_exception=True
)
def applicant_list(request):

    applicants = Applicant.objects.all()

    q = request.GET.get("q")

    if q:
        applicants = applicants.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(application_no__icontains=q)
        )

    return render(
        request,
        "students/applicants/applicant_list.html",
        {
            "applicants": applicants
        }
    )

@login_required
@permission_required(
    "students.add_applicant",
    raise_exception=True
)
def applicant_create(request):

    if request.method == "POST":

        form = ApplicantForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Applicant created."
            )

            return redirect(
                "applicant_list"
            )

    else:

        form = ApplicantForm()

    return render(
        request,
        "students/applicants/applicant_form.html",
        {
            "form": form
        }
    )

@login_required
@permission_required(
    "students.change_applicant",
    raise_exception=True
)
def applicant_update(
    request,
    pk
):

    applicant = get_object_or_404(
        Applicant,
        pk=pk
    )

    if request.method == "POST":

        form = ApplicantForm(
            request.POST,
            instance=applicant
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Applicant updated."
            )

            return redirect(
                "applicant_list"
            )

    else:

        form = ApplicantForm(
            instance=applicant
        )

    return render(
        request,
        "students/applicants/applicant_form.html",
        {
            "form": form
        }
    )


@login_required
@permission_required(
    "students.delete_applicant",
    raise_exception=True
)
def applicant_delete(
    request,
    pk
):

    applicant = get_object_or_404(
        Applicant,
        pk=pk
    )

    if request.method == "POST":

        applicant.delete()

        messages.success(
            request,
            "Applicant deleted."
        )

        return redirect(
            "applicant_list"
        )

    return render(
        request,
        "students/applicants/applicant_confirm_delete.html",
        {
            "applicant": applicant
        }
    )

@login_required
@permission_required(
    "students.change_applicant",
    raise_exception=True
)
def approve_applicant(request, pk):

    applicant = get_object_or_404(
        Applicant,
        pk=pk
    )


    if applicant.student:

        messages.warning(
            request,
            "Applicant has already been approved."
        )

        return redirect(
            "applicant_detail",
            pk=applicant.pk
        )


    active_year = (
        AcademicYear.objects
        .filter(
            is_active=True
        )
        .first()
    )


    active_semester = (
        Semester.objects
        .filter(
            is_active=True
        )
        .first()
    )


    if not active_year:

        messages.error(
            request,
            "No active academic year configured."
        )

        return redirect(
            "applicant_detail",
            pk=applicant.pk
        )


    if not active_semester:

        messages.error(
            request,
            "No active semester configured."
        )

        return redirect(
            "applicant_detail",
            pk=applicant.pk
        )


    # First semester of the programme

    first_level = (
        ProgrammeLevel.objects
        .filter(
            programme=applicant.programme,
            is_active=True
        )
        .order_by(
            "progression_order"
        )
        .first()
    )


    if not first_level:

        messages.error(
            request,
            "Programme levels have not been configured."
        )

        return redirect(
            "applicant_detail",
            pk=applicant.pk
        )


    try:

        with transaction.atomic():


            student = Student.objects.create(

                first_name=applicant.first_name,

                middle_name=applicant.middle_name,

                last_name=applicant.last_name,

                gender=applicant.gender,

                date_of_birth=applicant.date_of_birth,

                id_number=applicant.id_number,

                phone=applicant.phone_number,

                email=applicant.email,

                address=applicant.address,

                programme=applicant.programme,

                admission_date=timezone.now().date(),

            )


            applicant.student = student

            applicant.status = "APPROVED"

            applicant.save()



            SemesterEnrollment.objects.create(

                student=student,

                programme=applicant.programme,

                programme_level=first_level,

                academic_year=active_year,

                semester=active_semester,

                status=SemesterEnrollment.ENROLLED,

            )


        messages.success(
            request,
            (
                "Applicant approved successfully. "
                f"Student {student.admission_no} created."
            )
        )


    except Exception as e:

        messages.error(
            request,
            f"Approval failed: {e}"
        )


    return redirect(
        "applicant_detail",
        pk=applicant.pk
    )

@login_required
@permission_required(
    "students.view_applicant",
    raise_exception=True
)
def applicant_detail(request, pk):

    applicant = get_object_or_404(
        Applicant,
        pk=pk
    )

    return render(
        request,
        "students/applicants/applicant_detail.html",
        {
            "applicant": applicant
        }
    )


@login_required
@permission_required(
    "students.view_intake",
    raise_exception=True
)
def intake_list(request):

    intakes = Intake.objects.all()

    return render(
        request,
        "students/intakes/intake_list.html",
        {
            "intakes": intakes
        }
    )

@login_required
@permission_required(
    "students.add_intake",
    raise_exception=True
)
def intake_create(request):

    if request.method == "POST":

        form = IntakeForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Intake created successfully."
            )

            return redirect(
                "intake_list"
            )

    else:

        form = IntakeForm()


    return render(
        request,
        "students/intakes/intake_form.html",
        {
            "form": form,
            "title": "Add Intake"
        }
    )

@login_required
@permission_required(
    "students.change_intake",
    raise_exception=True
)
def intake_update(request, pk):

    intake = get_object_or_404(
        Intake,
        pk=pk
    )


    if request.method == "POST":

        form = IntakeForm(
            request.POST,
            instance=intake
        )


        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Intake updated successfully."
            )

            return redirect(
                "intake_list"
            )

    else:

        form = IntakeForm(
            instance=intake
        )


    return render(
        request,
        "students/intakes/intake_form.html",
        {
            "form": form,
            "title": "Edit Intake"
        }
    )


@login_required
@permission_required(
    "students.delete_intake",
    raise_exception=True
)
def intake_delete(request, pk):

    intake = get_object_or_404(
        Intake,
        pk=pk
    )


    if request.method == "POST":

        intake.delete()

        messages.success(
            request,
            "Intake deleted successfully."
        )

        return redirect(
            "intake_list"
        )


    return render(
        request,
        "students/intakes/intake_confirm_delete.html",
        {
            "intake": intake
        }
    )

#Create Report View
@login_required
def admissions_report(request):

    by_intake = (
        Applicant.objects
        .values("intake__name")
        .annotate(total=Count("id"))
        .order_by("intake__name")
    )

    by_programme = (
        Applicant.objects
        .values("programme__programme_name")
        .annotate(total=Count("id"))
        .order_by("programme__programme_name")
    )

    by_department = (
        Applicant.objects
        .values("programme__department__department_name")
        .annotate(total=Count("id"))
        .order_by("programme__department__department_name")
    )

    context = {
        "by_intake": by_intake,
        "by_programme": by_programme,
        "by_department": by_department,
    }

    return render(
        request,
        "students/admissions_report.html",
        context,
    )
"""
@login_required
def lecturer_assignment_list(request):

    assignments = LecturerAssignment.objects.select_related(
        "lecturer",
        "unit",
        "academic_year",
        "semester"
    )

    return render(
        request,
        "students/assignments/assignment_list.html",
        {
            "assignments": assignments
        }
    )

@login_required
def lecturer_assignment_create(request):

    if request.method == "POST":

        form = LecturerAssignmentForm(request.POST)

        if form.is_valid():

            assignment = form.save(commit=False)

            assignment.assigned_by = request.user

            assignment.save()

            return redirect(
                "lecturer_assignment_list"
            )

    else:

        form = LecturerAssignmentForm()


    return render(
        request,
        "students/assignments/assignment_form.html",
        {
            "form": form
        }
    )

"""
@login_required
@permission_required(
    "students.view_lecturerassignment",
    raise_exception=True
)
def lecturer_assignment_list(request):

    assignments = LecturerAssignment.objects.select_related(
        "lecturer",
        "unit_offering",
        "unit_offering__academic_year",
        "unit_offering__semester",
        "unit_offering__programme_level",
        "unit_offering__programme_level__programme",
        "unit_offering__unit",
    )

    return render(
        request,
        "students/assignments/assignment_list.html",
        {
            "assignments": assignments
        }
    )

@login_required
@permission_required(
    "students.add_lecturerassignment",
    raise_exception=True
)
def lecturer_assignment_create(request):

    if request.method == "POST":

        form = LecturerAssignmentForm(request.POST)

        if form.is_valid():

            assignment = form.save(commit=False)

            assignment.assigned_by = request.user

            assignment.save()

            return redirect(
                "lecturer_assignment_list"
            )

    else:

        form = LecturerAssignmentForm()

    return render(
        request,
        "students/assignments/assignment_form.html",
        {
            "form": form,
            "title": "Assign Unit to Lecturer",
        }
    )

@login_required
@permission_required(
    "students.change_lecturerassignment",
    raise_exception=True
)
def lecturer_assignment_update(request, pk):

    assignment = get_object_or_404(
        LecturerAssignment,
        pk=pk
    )

    if request.method == "POST":

        form = LecturerAssignmentForm(
            request.POST,
            instance=assignment
        )

        if form.is_valid():

            form.save()

            return redirect(
                "lecturer_assignment_list"
            )

    else:

        form = LecturerAssignmentForm(
            instance=assignment
        )

    return render(
        request,
        "students/assignments/assignment_form.html",
        {
            "form": form,
            "title": "Edit Lecturer Assignment",
        }
    )

@login_required
@permission_required(
    "students.delete_lecturerassignment",
    raise_exception=True
)
def lecturer_assignment_delete(request, pk):

    assignment = get_object_or_404(
        LecturerAssignment,
        pk=pk
    )

    if request.method == "POST":

        assignment.delete()

        return redirect(
            "lecturer_assignment_list"
        )

    return render(
        request,
        "students/assignments/assignment_confirm_delete.html",
        {
            "assignment": assignment
        }
    )

"""
@login_required
def my_units(request):

    assignments = LecturerAssignment.objects.select_related(
        "unit",
        "academic_year",
        "semester",
    ).filter(
        lecturer=request.user
    )

    return render(
        request,
        "students/exams/my_units.html",
        {
            "assignments": assignments
        }
    )"""


@login_required
def my_units(request):

    assignments = LecturerAssignment.objects.select_related(
        "unit_offering",
        "unit_offering__unit",
        "unit_offering__academic_year",
        "unit_offering__semester",
        "unit_offering__programme_level",
    ).filter(
        lecturer=request.user
    )

    assignment_data = []

    for assignment in assignments:

        batch = ResultBatch.objects.filter(
            lecturer_assignment=assignment
        ).first()

        registered_students = Registration.objects.filter(
            enrollment__academic_year=assignment.unit_offering.academic_year,
            enrollment__semester=assignment.unit_offering.semester,
            unit=assignment.unit_offering.unit,
            status=Registration.REGISTERED,
        ).count()


        assignment_data.append(
            {
                "assignment": assignment,
                "batch": batch,
                "registered_students": registered_students,
            }
        )


    return render(
        request,
        "students/exams/my_units.html",
        {
            "assignment_data": assignment_data,
        }
    )

from decimal import Decimal

@login_required
def enter_marks(request, assignment_id):

    assignment = get_object_or_404(
        LecturerAssignment,
        id=assignment_id,
        lecturer=request.user,
    )

    offering = assignment.unit_offering

    registrations = (
        Registration.objects
        .select_related(
            "enrollment",
            "enrollment__student",
            "unit",
        )
        .filter(
            enrollment__academic_year=offering.academic_year,
            enrollment__semester=offering.semester,
            unit=offering.unit,
            status=Registration.REGISTERED,
        )
        .order_by(
            "enrollment__student__admission_no"
        )
    )

    batch, created = ResultBatch.objects.get_or_create(
        lecturer_assignment=assignment,
        defaults={
            "unit_offering": offering,
        },
    )

    locked = batch.status in [
        ResultBatch.SUBMITTED,
        ResultBatch.APPROVED,
    ]

    # Load ONLY existing results
    result_map = {
        result.enrollment_id: result
        for result in Result.objects.filter(
            unit_offering=offering
        )
    }

    if request.method == "POST" and not locked:

        for registration in registrations:

            cat1 = request.POST.get(
                f"cat1_{registration.id}",
                "",
            ).strip()

            cat2 = request.POST.get(
                f"cat2_{registration.id}",
                "",
            ).strip()

            exam = request.POST.get(
                f"exam_{registration.id}",
                "",
            ).strip()

            # Ignore completely blank rows
            if not cat1 and not cat2 and not exam:
                continue

            result = result_map.get(
                registration.enrollment.id
            )

            if result is None:

                result = Result(
                    enrollment=registration.enrollment,
                    unit_offering=offering,
                )

            result.cat1 = Decimal(cat1) if cat1 else None
            result.cat2 = Decimal(cat2) if cat2 else None
            result.exam = Decimal(exam) if exam else None

            result.entered_by = request.user
            result.batch = batch

            result.save()

            result_map[registration.enrollment.id] = result

        messages.success(
            request,
            "Marks saved successfully.",
        )

        return redirect(
            "enter_marks",
            assignment_id=assignment.id,
        )

    context = {
        "assignment": assignment,
        "offering": offering,
        "registrations": registrations,
        "result_map": result_map,
        "locked": locked,
    }

    return render(
        request,
        "students/exams/enter_marks.html",
        context,
    )

@login_required
def unit_marksheet(request, assignment_id):

    assignment = get_object_or_404(
        LecturerAssignment,
        id=assignment_id,
    )

    offering = assignment.unit_offering

    batch = (
        ResultBatch.objects
        .filter(
            lecturer_assignment=assignment,
        )
        .first()
    )

    results = (
        Result.objects
        .select_related(
            "enrollment",
            "enrollment__student",
            "unit_offering",
            "unit_offering__unit",
            "entered_by",
        )
        .filter(
            unit_offering=offering,
            enrollment__registrations__unit=offering.unit,
            enrollment__registrations__status=Registration.REGISTERED,
        )
        .distinct()
        .order_by(
            "enrollment__student__admission_no",
        )
    )

    status = "Draft"
    examiner_remark = ""

    if batch:

        examiner_remark = batch.remarks or ""

        if batch.status == ResultBatch.DRAFT:
            status = "Draft"

        elif batch.status == ResultBatch.SUBMITTED:
            status = "Submitted"

        elif batch.status == ResultBatch.APPROVED:
            status = "Approved"

        elif batch.status == ResultBatch.RETURNED:
            status = "Returned"

        elif batch.status == ResultBatch.UNLOCKED:
            status = "Unlocked"

    return render(
        request,
        "students/exams/unit_marksheet.html",
        {
            "assignment": assignment,
            "offering": offering,
            "batch": batch,
            "results": results,
            "status": status,
            "examiner_remark": examiner_remark,
        },
    )

@login_required
def export_marksheet_excel(request, assignment_id):

    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        Alignment,
        Border,
        Side,
    )
    from openpyxl.utils import get_column_letter

    assignment = get_object_or_404(
        LecturerAssignment,
        id=assignment_id,
    )

    offering = assignment.unit_offering

    batch = (
        ResultBatch.objects
        .filter(
            lecturer_assignment=assignment,
            unit_offering=offering,
        )
        .first()
    )

    results = (
        Result.objects
        .select_related(
            "enrollment",
            "enrollment__student",
            "unit_offering",
            "unit_offering__unit",
            "unit_offering__programme_level",
        )
        .filter(
            unit_offering=offering,
            enrollment__registrations__unit=offering.unit,
            enrollment__registrations__status=Registration.REGISTERED,
        )
        .distinct()
        .order_by(
            "enrollment__student__admission_no",
        )
    )

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Marksheet"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ==========================
    # HEADER
    # ==========================

    sheet.merge_cells("A1:H1")
    sheet["A1"] = "SANFIELDS INSTITUTE OF BUSINESS & TECHNOLOGY"
    sheet["A1"].font = Font(
        bold=True,
        size=16,
    )
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    sheet.merge_cells("A2:H2")
    sheet["A2"] = "OFFICIAL UNIT MARKSHEET"
    sheet["A2"].font = Font(
        bold=True,
        size=12,
    )
    sheet["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    sheet["A3"] = "Workflow Status"
    sheet["B3"] = batch.get_status_display() if batch else "Draft"

    sheet["E3"] = "Generated"
    sheet["F3"] = timezone.now().strftime("%d %b %Y %H:%M")

    sheet["A4"] = "Unit"
    sheet["B4"] = f"{offering.unit.code} - {offering.unit.name}"

    sheet["E4"] = "Programme Level"
    sheet["F4"] = str(offering.programme_level)

    sheet["A5"] = "Programme"
    sheet["B5"] = str(offering.programme_level.programme)

    sheet["E5"] = "Academic Year"
    sheet["F5"] = str(offering.academic_year)

    sheet["A6"] = "Semester"
    sheet["B6"] = str(offering.semester)

    sheet["E6"] = "Lecturer"
    sheet["F6"] = (
        assignment.lecturer.get_full_name()
        or assignment.lecturer.username
    )

    if batch and batch.remarks:
        sheet.merge_cells("A7:H7")
        sheet["A7"] = (
            f"Exam Officer Remarks: {batch.remarks}"
        )
        sheet["A7"].font = Font(
            italic=True,
            color="AA0000",
        )

    # ==========================
    # TABLE HEADERS
    # ==========================

    headers = [
        "Admission No",
        "Student Name",
        "CAT 1",
        "CAT 2",
        "Exam",
        "Total",
        "Grade",
        "Remarks",
    ]

    header_row = 9

    for col_num, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=header_row,
            column=col_num,
        )

        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border

    # ==========================
    # RESULTS
    # ==========================

    row_num = 10

    for result in results:

        values = [
            result.enrollment.student.admission_no,
            str(result.enrollment.student),
            result.cat1,
            result.cat2,
            result.exam,
            result.total,
            result.grade,
            result.remarks,
        ]

        for col_num, value in enumerate(values, start=1):

            cell = sheet.cell(
                row=row_num,
                column=col_num,
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        row_num += 1

    # ==========================
    # AUTO WIDTH
    # ==========================

    for col in range(1, 9):

        letter = get_column_letter(col)
        longest = 0

        for row in range(1, sheet.max_row + 1):

            value = sheet.cell(
                row=row,
                column=col,
            ).value

            if value:
                longest = max(
                    longest,
                    len(str(value)),
                )

        sheet.column_dimensions[letter].width = longest + 5

    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["H"].width = 25

    # ==========================
    # SIGNATURES
    # ==========================

    row_num += 3

    sheet.cell(
        row=row_num,
        column=1,
    ).value = "Lecturer Signature"

    sheet.cell(
        row=row_num,
        column=4,
    ).value = "HOD Signature"

    sheet.cell(
        row=row_num,
        column=7,
    ).value = "Exam Officer Signature"

    # ==========================
    # DOWNLOAD
    # ==========================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{offering.unit.code}_marksheet.xlsx"'
    )

    workbook.save(response)

    return response

@login_required
def submit_results(request, assignment_id):

    assignment = get_object_or_404(
        LecturerAssignment,
        id=assignment_id,
        lecturer=request.user,
    )

    offering = assignment.unit_offering

    batch, created = ResultBatch.objects.get_or_create(
        lecturer_assignment=assignment,
        unit_offering=offering,
    )

    # Prevent duplicate submissions

    if batch.status == ResultBatch.SUBMITTED:

        messages.warning(
            request,
            "These results have already been submitted and are awaiting approval.",
        )

        return redirect(
            "enter_marks",
            assignment_id=assignment.id,
        )

    if batch.status == ResultBatch.APPROVED:

        messages.error(
            request,
            "These results have already been approved.",
        )

        return redirect(
            "enter_marks",
            assignment_id=assignment.id,
        )

    # Get all results for this offering

    results = Result.objects.filter(
        unit_offering=offering,
    )

    # Ensure result records exist

    registered_count = Registration.objects.filter(
        enrollment__academic_year=offering.academic_year,
        enrollment__semester=offering.semester,
        unit=offering.unit,
        status=Registration.REGISTERED,
    ).count()

    if results.count() < registered_count:

        messages.error(
            request,
            "Some registered students do not yet have result records."
        )

        return redirect(
            "enter_marks",
            assignment_id=assignment.id,
        )

    # Attach all results to this batch

    results.update(
        batch=batch,
    )

    # Submit batch regardless of blank marks

    batch.status = ResultBatch.SUBMITTED
    batch.submitted_by = request.user
    batch.submitted_at = timezone.now()
    batch.approved_by = None
    batch.approved_at = None
    batch.remarks = ""
    batch.save()

    blank_results = results.filter(
        cat1__isnull=True,
        cat2__isnull=True,
        exam__isnull=True,
    ).count()

    if blank_results:

        messages.warning(
            request,
            f"Results submitted successfully. "
            f"{blank_results} student(s) have no marks entered. "
            f"The Examination Office should verify before approval."
        )

    else:

        messages.success(
            request,
            "Results submitted successfully and forwarded to the Examination Office."
        )

    return redirect(
        "my_units",
    )

@login_required
@permission_required(
    "students.view_resultbatch",
    raise_exception=True
)
def approval_queue(request):

    batches = ResultBatch.objects.select_related(
        "unit_offering",
        "unit_offering__unit",
        "unit_offering__academic_year",
        "unit_offering__semester",
        "submitted_by",
        "approved_by",
    ).exclude(
        status="draft"
    )


    return render(
        request,
        "students/exams/approval_queue.html",
        {
            "batches": batches
        }
    )

@login_required
@permission_required(
    "students.change_resultbatch",
    raise_exception=True
)
def approve_results(request, batch_id):

    batch = get_object_or_404(
        ResultBatch,
        id=batch_id
    )

    # Prevent approving already approved batches

    if batch.status == "approved":

        messages.warning(
            request,
            "This result batch is already approved."
        )

        return redirect(
            "approval_queue"
        )


    # Only submitted batches should be approved

    if batch.status != "submitted":

        messages.error(
            request,
            "Only submitted result batches can be approved."
        )

        return redirect(
            "approval_queue"
        )


    batch.status = "approved"

    batch.approved_by = request.user

    batch.approved_at = timezone.now()

    batch.save()


    messages.success(
        request,
        "Results approved successfully."
    )


    return redirect(
        "approval_queue"
    )


@login_required
@permission_required(
    "students.change_resultbatch",
    raise_exception=True
)


@login_required
@transaction.atomic
def publish_results(request, batch_id):

    batch = get_object_or_404(
        ResultBatch,
        pk=batch_id,
    )

    if batch.status != ResultBatch.APPROVED:

        messages.error(
            request,
            "Only approved result batches can be published."
        )

        return redirect(
            "approval_queue"
        )

    batch.status = ResultBatch.PUBLISHED

    batch.published_by = request.user

    batch.published_at = timezone.now()

    batch.save()

    ResultBatchLog.objects.create(
        batch=batch,
        action="Published",
        performed_by=request.user,
        remarks="Results published for student access.",
    )

    messages.success(
        request,
        "Results published successfully."
    )

    return redirect(
        "approval_queue"
    )

def return_results(request, batch_id):

    batch = get_object_or_404(
        ResultBatch,
        id=batch_id
    )

    if request.method == "POST":

        remarks = request.POST.get(
            "remarks"
        )

        if not remarks:

            messages.error(
                request,
                "Please provide a reason for returning results."
            )

            return render(
                request,
                "students/exams/return_results.html",
                {
                    "batch": batch
                }
            )

        batch.status = "returned"

        batch.remarks = remarks

        batch.save()

        messages.success(
            request,
            "Results returned for correction."
        )

        return redirect(
            "approval_queue"
        )


    return render(
        request,
        "students/exams/return_results.html",
        {
            "batch": batch
        }
    )


@login_required
@permission_required(
    "students.view_resultbatch",
    raise_exception=True
)
def view_batch(request, batch_id):

    batch = get_object_or_404(
        ResultBatch,
        id=batch_id
    )

    results = Result.objects.select_related(
        "enrollment__student"
    ).filter(
        batch=batch
    )

    return render(
        request,
        "students/exams/view_batch.html",
        {
            "batch": batch,
            "results": results,
        }
    )

@login_required
@permission_required(
    "students.view_resultbatch",
    raise_exception=True,
)
def batch_details(request, batch_id):

    batch = get_object_or_404(
        ResultBatch.objects.select_related(
            "unit_offering",
            "unit_offering__unit",
            "unit_offering__academic_year",
            "unit_offering__semester",
            "unit_offering__programme_level",
            "lecturer_assignment",
            "lecturer_assignment__lecturer",
            "submitted_by",
            "approved_by",
        ),
        id=batch_id,
    )

    offering = batch.unit_offering

    results = (
        Result.objects.select_related(
            "enrollment__student",
            "unit_offering",
            "unit_offering__unit",
        )
        .filter(
            batch=batch,
        )
        .order_by(
            "enrollment__student__admission_no",
        )
    )

    registered_students = Registration.objects.filter(
        unit=batch.unit_offering.unit,
        enrollment__academic_year=batch.unit_offering.academic_year,
        enrollment__semester=batch.unit_offering.semester,
        status=Registration.REGISTERED,
    ).count()

    return render(
        request,
        "students/exams/batch_details.html",
        {
            "batch": batch,
            "offering": offering,
            "results": results,
            "registered_students": registered_students,
            "submitted_students": results.count(),
        },
    )


@login_required
@permission_required(
    "students.change_resultbatch",
    raise_exception=True
)
def unlock_batch(request, batch_id):

    batch = get_object_or_404(
        ResultBatch,
        id=batch_id
    )


    if request.method == "POST":

        remarks = request.POST.get(
            "remarks"
        )


        if not remarks:

            messages.error(
                request,
                "Unlock reason is required."
            )

            return render(
                request,
                "students/exams/unlock_batch.html",
                {
                    "batch": batch
                }
            )


        batch.status = "unlocked"

        batch.remarks = remarks

        batch.save()


        ResultBatchLog.objects.create(
            batch=batch,
            action="Unlocked",
            performed_by=request.user,
            remarks=remarks,
        )


        messages.success(
            request,
            "Result batch unlocked successfully."
        )


        return redirect(
            "approval_queue"
        )


    return render(
        request,
        "students/exams/unlock_batch.html",
        {
            "batch": batch
        }
    )

@login_required
def student_results(request):

    student = get_object_or_404(
        Student,
        user=request.user,
    )

    enrollments = (
        SemesterEnrollment.objects
        .filter(student=student)
        .select_related(
            "academic_year",
            "semester",
            "programme_level",
        )
        .order_by(
            "academic_year",
            "semester",
        )
    )

    results = (
        Result.objects
        .select_related(
            "batch",
            "unit_offering",
            "unit_offering__unit",
            "unit_offering__programme_level",
            "enrollment",
            "enrollment__academic_year",
            "enrollment__semester",
        )
        .filter(
            enrollment__student=student,
            batch__status=ResultBatch.PUBLISHED,
            enrollment__registrations__unit=models.F(
                "unit_offering__unit"
            ),
            enrollment__registrations__status=Registration.REGISTERED,
        )
        .distinct()
        .order_by(
            "enrollment__academic_year",
            "enrollment__semester",
            "unit_offering__unit__code",
        )
    )

    return render(
        request,
        "students/results/student_results.html",
        {
            "student": student,
            "results": results,
            "enrollments": enrollments,
        },
    )

@login_required
def enrollment_results(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment.objects.select_related(
            "student",
            "programme_level",
            "programme_level__programme",
            "academic_year",
            "semester",
        ),
        pk=pk,
    )

    results = (
        Result.objects.filter(
            enrollment=enrollment,
            batch__status=ResultBatch.APPROVED,
        )
        .select_related(
            "batch",
            "unit_offering",
            "unit_offering__unit",
            "unit_offering__academic_year",
            "unit_offering__semester",
            "unit_offering__programme_level",
        )
        .order_by(
            "unit_offering__unit__code",
        )
    )

    return render(
        request,
        "students/results/enrollment_results.html",
        {
            "enrollment": enrollment,
            "student": enrollment.student,
            "results": results,
        },
    )


@login_required
@permission_required(
    "students.change_semesterenrollment",
    raise_exception=True
)
def progress_student(request, enrollment_id):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        id=enrollment_id
    )

    try:

        result = progress_student_service(
            enrollment=enrollment,
            user=request.user,
        )

        if result == enrollment:

            messages.success(
                request,
                "Student has completed the programme successfully."
            )

        else:

            messages.success(
                request,
                (
                    "Student progressed successfully to "
                    f"{result.programme_level.name}."
                )
            )

    except Exception as e:

        messages.error(
            request,
            str(e)
        )

    return redirect("enrollment_list")


@login_required
@permission_required(
    "students.view_semesterenrollment",
    raise_exception=True
)
def progression_list(request):

    enrollments = (
        SemesterEnrollment.objects
        .filter(
            status=SemesterEnrollment.ENROLLED
        )
        .select_related(
            "student",
            "programme",
            "programme_level",
            "academic_year",
            "semester",
        )
        .order_by(
            "programme",
            "programme_level__progression_order",
            "student__admission_no"
        )
    )


    progression_data = []


    for enrollment in enrollments:


        next_level = (
            ProgrammeLevel.objects
            .filter(
                programme=enrollment.programme,
                progression_order__gt=
                enrollment.programme_level.progression_order,
                is_active=True
            )
            .order_by(
                "progression_order"
            )
            .first()
        )


        progression_data.append(
            {
                "enrollment": enrollment,

                "next_level": next_level
            }
        )


    return render(
        request,
        "students/progression/progression_list.html",
        {
            "progression_data":
            progression_data
        }
    )


@login_required
@permission_required(
    "students.view_programmelevel",
    raise_exception=True
)
def programme_level_list(request):

    programme_levels = (
        ProgrammeLevel.objects
        .select_related(
            "programme",
            "programme__course",
            "programme__course__department",
        )
        .order_by(
            "programme__name",
            "progression_order",
        )
    )

    context = {
        "programme_levels": programme_levels
    }

    return render(
        request,
        "students/programme_levels/programme_level_list.html",
        context
    )

@login_required
@permission_required(
    "students.add_programmelevel",
    raise_exception=True
)
def programme_level_create(request):

    if request.method == "POST":

        form = ProgrammeLevelForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Programme Level created successfully."
            )

            return redirect(
                "programme_level_list"
            )

    else:

        form = ProgrammeLevelForm()


    return render(
        request,
        "students/programme_levels/programme_level_form.html",
        {
            "form": form
        }
    )


@login_required
@permission_required(
    "students.change_programmelevel",
    raise_exception=True
)
def programme_level_update(request, pk):

    level = get_object_or_404(
        ProgrammeLevel,
        pk=pk
    )


    if request.method == "POST":

        form = ProgrammeLevelForm(
            request.POST,
            instance=level
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Programme Level updated successfully."
            )

            return redirect(
                "programme_level_list"
            )


    else:

        form = ProgrammeLevelForm(
            instance=level
        )


    return render(
        request,
        "students/programme_levels/programme_level_form.html",
        {
            "form": form,
            "level": level,
        }
    )

@login_required
@permission_required(
    "students.delete_programmelevel",
    raise_exception=True
)
def programme_level_delete(request, pk):

    level = get_object_or_404(
        ProgrammeLevel,
        pk=pk
    )


    if request.method == "POST":

        level.delete()

        messages.success(
            request,
            "Programme Level deleted"
        )

        return redirect(
            "programme_level_list"
        )


    return render(
        request,
        "students/programme_levels/programme_level_confirm_delete.html",
        {
            "level": level
        }
    )


@login_required
def load_programme_levels(request):

    programme_id = request.GET.get(
        "programme_id"
    )


    levels = ProgrammeLevel.objects.filter(
        programme_id=programme_id
    ).order_by(
        "progression_order"
    )


    return JsonResponse(
        list(
            levels.values(
                "id",
                "name"
            )
        ),
        safe=False
    )

@login_required
def programme_level_units(request, pk):

    programme_level = get_object_or_404(
        ProgrammeLevel,
        pk=pk
    )

    units = programme_level.units.filter(
        is_active=True
    )

    context = {
        "programme_level": programme_level,
        "units": units,
    }

    return render(
        request,
        "students/programme_levels/programme_level_units.html",
        context
    )


def add_programme_level_unit(request, pk):

    programme_level = get_object_or_404(
        ProgrammeLevel,
        pk=pk
    )

    if request.method == "POST":

        code = request.POST.get("code")
        name = request.POST.get("name")
        credit_hours = request.POST.get("credit_hours")


        Unit.objects.create(

            programme_level=programme_level,

            code=code,

            name=name,

            credit_hours=credit_hours

        )


        return redirect(
            "programme_level_units",
            pk=programme_level.id
        )


    context = {
        "programme_level": programme_level
    }


    return render(
        request,
        "students/programme_levels/add_unit.html",
        context
    )

def semester_enrollment_list(request):

    enrollments = SemesterEnrollment.objects.select_related(
        "student",
        "programme",
        "programme_level",
        "academic_year",
        "semester",
    )

    return render(
        request,
        "students/enrollments/enrollment_list.html",
        {
            "enrollments": enrollments
        }
    )

@transaction.atomic
def semester_enrollment_create(request):

    form = SemesterEnrollmentForm(
        request.POST or None
    )


    if form.is_valid():


        enrollment = form.save()



        try:

            generate_student_invoice(
                enrollment
            )


            update_financial_clearance(
                enrollment,
                request.user
            )


            messages.success(

                request,

                "Semester enrollment created and invoice generated successfully."

            )


        except ValueError as e:


            messages.warning(

                request,

                str(e)

            )



        return redirect(

            "semester_enrollment_detail",

            pk=enrollment.pk

        )



    return render(

        request,

        "students/enrollments/enrollment_form.html",

        {

            "form": form

        }

    )


def semester_enrollment_detail(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )

    registrations = enrollment.registrations.all()


    return render(
        request,
        "students/enrollments/enrollment_detail.html",
        {
            "enrollment": enrollment,
            "registrations": registrations
        }
    )

def semester_enrollment_edit(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )


    form = SemesterEnrollmentForm(
        request.POST or None,
        instance=enrollment
    )


    if form.is_valid():

        form.save()

        return redirect(
            "semester_enrollment_list"
        )


    return render(
        request,
        "students/enrollments/enrollment_form.html",
        {
            "form":form
        }
    )


def semester_enrollment_delete(request, pk):

    enrollment = get_object_or_404(
        SemesterEnrollment,
        pk=pk
    )


    if request.method == "POST":

        enrollment.delete()

        return redirect(
            "semester_enrollment_list"
        )


    return render(
        request,
        "students/enrollments/enrollment_confirm_delete.html",
        {
            "enrollment": enrollment
        }
    )


@login_required
@transaction.atomic
def enroll_student(request, pk):
    student = get_object_or_404(
        Student.objects.select_related(
            "programme"
        ),
        pk=pk

    )
    academic_year = AcademicYear.objects.filter(

        is_active=True

    ).first()

    if not academic_year:
        messages.error(

            request,
            "There is no active academic year."
        )
        return redirect(

            "student_detail",

            id=student.pk

        )

    semester = Semester.objects.filter(
        is_active=True

    ).first()

    if not semester:
        messages.error(
            request,
            "There is no active semester."
        )
        return redirect(

            "student_detail",

            id=student.pk

        )
    programme_level = ProgrammeLevel.objects.filter(

        programme=student.programme,

        is_active=True,

    ).order_by(

        "progression_order"

    ).first()

    if not programme_level:
        messages.error(
            request,

            "No active programme level found."
        )
        return redirect(

            "student_detail",

            id=student.pk

        )
    enrollment, created = SemesterEnrollment.objects.get_or_create(
        student=student,
        academic_year=academic_year,
        semester=semester,
        defaults={
            "programme": student.programme,

            "programme_level": programme_level,

            "status": SemesterEnrollment.ENROLLED,
        }
    )

    if created:
        try:
            generate_student_invoice(
                enrollment
            )

            update_financial_clearance(
                enrollment,
                request.user
            )

            messages.success(
                request,
                "Student enrolled and invoice generated successfully."
            )

        except ValueError as e:
            messages.warning(
                request,
                str(e)

            )

    else:
        messages.info(
            request,
            "Student is already enrolled for the active semester."
        )
    return redirect(
        "semester_enrollment_detail",
        pk=enrollment.pk

    )

class StudentDeleteView(DeleteView):

    model = Student

    template_name = "students/students/student_delete.html"

    context_object_name = "student"

    success_url = reverse_lazy("student_list")

    def post(self, request, *args, **kwargs):

        self.object = self.get_object()

        if self.object.enrollments.exists():

            messages.error(

                request,

                (
                    "This student cannot be deleted because "
                    "semester enrollment records exist."
                ),

            )

            return redirect(

                "student_detail",

                self.object.pk,

            )

        return super().post(request, *args, **kwargs)